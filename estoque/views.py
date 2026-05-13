from django.shortcuts import render, redirect, get_object_or_404
from decimal import Decimal
from functools import wraps
import csv
from io import BytesIO

from django.db.models import Count, F, Sum
from django.db.models.deletion import ProtectedError
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from .models import (
    Produto,
    Pedido,
    Fornecedor,
    Unidade,
    Setor,
    Entrada,
    ItemEntrada,
    ItemPedido,
    PerfilUsuario,
)
from .forms import (
    ProdutoForm,
    FornecedorForm,
    UnidadeForm,
    SetorForm,
    EntradaForm,
    ItemEntradaFormSet,
    PedidoForm,
    ItemPedidoFormSet,
    AnexoEmpenhoForm,
    PerfilUsuarioForm,
    PerfilTemaForm,
)


def _tem_papel(*nomes_grupos):
    """Retorna True se o usuário pertence a um dos grupos ou é superusuário."""
    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def _wrapped(request, *args, **kwargs):
            if request.user.is_superuser or request.user.groups.filter(name__in=nomes_grupos).exists():
                return view_func(request, *args, **kwargs)
            messages.error(request, "Você não tem permissão para acessar esta funcionalidade.")
            return redirect("dashboard")
        return _wrapped
    return decorator



def _to_decimal(value):
    return value if value is not None else Decimal("0")


@login_required
def produtos_por_fornecedor(request):
    fornecedor_id = request.GET.get("fornecedor_id")
    if not fornecedor_id:
        return JsonResponse({"produtos": []})

    produtos = Produto.objects.filter(fornecedores__id=fornecedor_id).order_by("nome")

    payload = [
        {
            "id": produto.id,
            "nome": produto.nome,
            "unidade_medida": produto.unidade_medida,
        }
        for produto in produtos
    ]
    return JsonResponse({"produtos": payload})


@login_required
def setores_por_unidade(request):
    unidade_id = request.GET.get("unidade_id")
    if not unidade_id:
        return JsonResponse({"setores": []})

    setores = Setor.objects.filter(unidade_id=unidade_id).order_by("nome")
    payload = [{"id": setor.id, "nome": setor.nome} for setor in setores]
    return JsonResponse({"setores": payload})


@login_required
def dashboard(request):
    total_produtos = Produto.objects.count()
    pedidos_reservados = Pedido.objects.filter(status="RESERVADO").count()
    entregas_realizadas = Pedido.objects.filter(status="ENTREGUE").count()

    # Produtos com estoque abaixo do mínimo
    estoque_critico = Produto.objects.filter(estoque_atual__lt=F("estoque_minimo"))

    # Pedidos recentes
    pedidos_recentes = (
        Pedido.objects.select_related("secretaria").all().order_by("-data_pedido")[:5]
    )

    entradas_recentes = Entrada.objects.select_related("fornecedor").order_by(
        "-data_entrada"
    )[:5]
    pedidos_pendentes_empenho = (
        Pedido.objects.select_related("secretaria")
        .filter(status="RESERVADO")
        .order_by("-data_pedido")[:6]
    )

    context = {
        "total_produtos": total_produtos,
        "pedidos_reservados": pedidos_reservados,
        "entregas_realizadas": entregas_realizadas,
        "estoque_critico": estoque_critico,
        "pedidos_recentes": pedidos_recentes,
        "entradas_recentes": entradas_recentes,
        "pedidos_pendentes_empenho": pedidos_pendentes_empenho,
    }
    return render(request, "estoque/dashboard.html", context)


# --- PRODUTOS ---
@login_required
def produto_list(request):
    search = request.GET.get("search", "")
    produtos = Produto.objects.all()
    if search:
        produtos = produtos.filter(nome__icontains=search)
    return render(
        request, "estoque/produto_list.html", {"produtos": produtos, "search": search}
    )


@_tem_papel("Almoxarife", "Administrador")
def produto_create(request):
    form = ProdutoForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Produto cadastrado com sucesso!")
        return redirect("produto_list")
    return render(
        request, "estoque/produto_form.html", {"form": form, "title": "Novo Produto"}
    )


@_tem_papel("Almoxarife", "Administrador")
def produto_update(request, pk):
    produto = get_object_or_404(Produto, pk=pk)
    form = ProdutoForm(request.POST or None, instance=produto)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Produto atualizado com sucesso!")
        return redirect("produto_list")
    return render(
        request, "estoque/produto_form.html", {"form": form, "title": "Editar Produto"}
    )


# --- FORNECEDORES ---
@login_required
def fornecedor_list(request):
    fornecedores = Fornecedor.objects.all()
    return render(
        request, "estoque/fornecedor_list.html", {"fornecedores": fornecedores}
    )


@_tem_papel("Almoxarife", "Comprador", "Administrador")
def fornecedor_create(request):
    form = FornecedorForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Fornecedor cadastrado com sucesso!")
        return redirect("fornecedor_list")
    return render(
        request,
        "estoque/fornecedor_form.html",
        {"form": form, "title": "Novo Fornecedor"},
    )


# --- UNIDADES ---
@login_required
def unidade_list(request):
    unidades = Unidade.objects.all().order_by("nome")
    return render(request, "estoque/unidade_list.html", {"unidades": unidades})


@_tem_papel("Almoxarife", "Administrador")
def unidade_create(request):
    form = UnidadeForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Unidade cadastrada com sucesso!")
        return redirect("unidade_list")
    return render(
        request,
        "estoque/unidade_form.html",
        {"form": form, "title": "Nova Unidade"},
    )


@_tem_papel("Almoxarife", "Administrador")
def unidade_update(request, pk):
    unidade = get_object_or_404(Unidade, pk=pk)
    form = UnidadeForm(request.POST or None, instance=unidade)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Unidade atualizada com sucesso!")
        return redirect("unidade_list")
    return render(
        request,
        "estoque/unidade_form.html",
        {"form": form, "title": "Editar Unidade", "unidade": unidade},
    )


@_tem_papel("Almoxarife", "Administrador")
def unidade_delete(request, pk):
    unidade = get_object_or_404(Unidade, pk=pk)
    if request.method == "POST":
        try:
            unidade.delete()
            messages.success(request, "Unidade excluida com sucesso!")
        except ProtectedError:
            messages.error(
                request,
                "Nao foi possivel excluir a unidade porque ela esta vinculada a entradas, pedidos ou setores em uso.",
            )
    return redirect("unidade_list")


# --- SETORES ---
@login_required
def setor_list(request):
    setores = (
        Setor.objects.select_related("unidade").all().order_by("unidade__nome", "nome")
    )
    return render(request, "estoque/setor_list.html", {"setores": setores})


@_tem_papel("Almoxarife", "Administrador")
def setor_create(request):
    form = SetorForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Setor cadastrado com sucesso!")
        return redirect("setor_list")
    return render(
        request,
        "estoque/setor_form.html",
        {"form": form, "title": "Novo Setor"},
    )


@_tem_papel("Almoxarife", "Administrador")
def setor_update(request, pk):
    setor = get_object_or_404(Setor, pk=pk)
    form = SetorForm(request.POST or None, instance=setor)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Setor atualizado com sucesso!")
        return redirect("setor_list")
    return render(
        request,
        "estoque/setor_form.html",
        {"form": form, "title": "Editar Setor", "setor": setor},
    )


@_tem_papel("Almoxarife", "Administrador")
def setor_delete(request, pk):
    setor = get_object_or_404(Setor, pk=pk)
    if request.method == "POST":
        try:
            setor.delete()
            messages.success(request, "Setor excluido com sucesso!")
        except ProtectedError:
            messages.error(
                request,
                "Nao foi possivel excluir o setor porque ele esta vinculado a entradas ou pedidos.",
            )
    return redirect("setor_list")


# --- ENTRADAS ---
@login_required
def entrada_list(request):
    entradas = (
        Entrada.objects.select_related("fornecedor", "unidade", "setor")
        .all()
        .order_by("-data_entrada", "-id")
    )
    return render(request, "estoque/entrada_list.html", {"entradas": entradas})


@_tem_papel("Almoxarife", "Administrador")
def entrada_create(request):
    entrada_form = EntradaForm(request.POST or None)
    formset = ItemEntradaFormSet(request.POST or None)
    if request.method == "POST":
        if entrada_form.is_valid() and formset.is_valid():
            entrada = entrada_form.save()
            formset.instance = entrada
            formset.save()
            messages.success(request, "Entrada de estoque registrada com sucesso!")
            return redirect("entrada_list")
    return render(
        request,
        "estoque/entrada_form.html",
        {"entrada_form": entrada_form, "formset": formset},
    )


# --- PEDIDOS ---
@login_required
def pedido_list(request):
    status_filter = request.GET.get("status")
    pedidos = (
        Pedido.objects.select_related("secretaria", "setor")
        .all()
        .order_by("-data_pedido")
    )
    if status_filter:
        pedidos = pedidos.filter(status=status_filter)
    return render(
        request,
        "estoque/pedido_list.html",
        {
            "pedidos": pedidos,
            "status_filter": status_filter,
        },
    )


@_tem_papel("Almoxarife", "Comprador", "Solicitante", "Administrador")
def pedido_create(request):
    pedido_form = PedidoForm(request.POST or None)
    formset = ItemPedidoFormSet(request.POST or None)
    if request.method == "POST":
        if pedido_form.is_valid() and formset.is_valid():
            pedido = pedido_form.save()
            formset.instance = pedido
            formset.save()
            messages.success(request, "Pedido solicitado com sucesso!")
            return redirect("pedido_list")
    return render(
        request,
        "estoque/pedido_form.html",
        {"pedido_form": pedido_form, "formset": formset},
    )


@login_required
def pedido_detail(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    empenho_form = AnexoEmpenhoForm(
        request.POST or None, request.FILES or None, instance=pedido
    )

    itens = list(pedido.itens.select_related("produto", "produto__categoria"))

    total_geral = Decimal("0")
    total_atendido = Decimal("0")
    total_licitado = Decimal("0")
    total_restante = Decimal("0")
    for item in itens:
        total_geral += _to_decimal(item.total_pedido)
        total_atendido += _to_decimal(item.total_atendido)
        total_licitado += _to_decimal(item.quantidade_licitada)
        total_restante += _to_decimal(item.restante_licitacao)

    if request.method == "POST":
        if "reservar" in request.POST:
            try:
                pedido.status = "RESERVADO"
                pedido.save()
                messages.success(request, "Pedido reservado com sucesso!")
            except ValueError as e:
                messages.error(request, str(e))

        elif "anexar_empenho" in request.POST and empenho_form.is_valid():
            empenho_form.save()
            pedido.status = "EMPENHADO"
            pedido.save()
            messages.success(request, "Empenho anexado e status atualizado!")

        elif "confirmar_entrega" in request.POST:
            pedido.status = "ENTREGUE"
            pedido.save()
            messages.success(request, "Entrega confirmada e estoque baixado!")

        return redirect("pedido_detail", pk=pk)

    return render(
        request,
        "estoque/pedido_detail.html",
        {
            "pedido": pedido,
            "itens": itens,
            "empenho_form": empenho_form,
            "total_geral": total_geral,
            "total_atendido": total_atendido,
            "total_licitado": total_licitado,
            "total_restante": total_restante,
        },
    )


# --- RELATÓRIOS ---
def _export_csv(filename, headers, rows):
    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


def _export_xlsx(filename, sheets):
    from openpyxl import Workbook
    wb = Workbook()
    for i, (sheet_name, headers, rows) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet(title=sheet_name)
        if i == 0:
            ws.title = sheet_name
        ws.append(headers)
        for row in rows:
            ws.append([str(v) if v is not None else "" for v in row])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def relatorios(request):
    return redirect("relatorio_movimento")


@login_required
def relatorio_movimento(request):
    data_de = request.GET.get("data_de", "")
    data_ate = request.GET.get("data_ate", "")
    secretaria_id = request.GET.get("secretaria_id", "")

    consumo_qs = ItemPedido.objects.filter(pedido__status="ENTREGUE")
    if data_de:
        consumo_qs = consumo_qs.filter(pedido__data_pedido__date__gte=data_de)
    if data_ate:
        consumo_qs = consumo_qs.filter(pedido__data_pedido__date__lte=data_ate)
    if secretaria_id:
        consumo_qs = consumo_qs.filter(pedido__secretaria_id=secretaria_id)

    consumo_secretaria = (
        consumo_qs
        .values("pedido__secretaria__nome")
        .annotate(total_itens=Sum("quantidade"))
        .order_by("-total_itens")
    )

    entradas_qs = ItemEntrada.objects.all()
    if data_de:
        entradas_qs = entradas_qs.filter(entrada__data_entrada__gte=data_de)
    if data_ate:
        entradas_qs = entradas_qs.filter(entrada__data_entrada__lte=data_ate)

    entradas_por_grupo = (
        entradas_qs
        .values("produto__categoria__nome")
        .annotate(total_quantidade=Sum("quantidade"))
        .order_by("-total_quantidade")
    )

    export = request.GET.get("export", "")
    if export == "csv":
        rows = [[r["pedido__secretaria__nome"] or "Sem secretaria", r["total_itens"]] for r in consumo_secretaria]
        return _export_csv("consumo_secretaria.csv", ["Secretaria", "Total Itens"], rows)
    if export == "xlsx":
        rows_consumo = [[r["pedido__secretaria__nome"] or "Sem secretaria", r["total_itens"]] for r in consumo_secretaria]
        rows_entradas = [[r["produto__categoria__nome"] or "Sem categoria", r["total_quantidade"]] for r in entradas_por_grupo]
        return _export_xlsx("relatorio_movimento.xlsx", [
            ("Consumo por Secretaria", ["Secretaria", "Total Itens"], rows_consumo),
            ("Entradas por Categoria", ["Categoria", "Total Quantidade"], rows_entradas),
        ])

    unidades = Unidade.objects.order_by("nome")
    return render(
        request,
        "estoque/relatorios.html",
        {
            "consumo_secretaria": consumo_secretaria,
            "entradas_por_grupo": entradas_por_grupo,
            "data_de": data_de,
            "data_ate": data_ate,
            "secretaria_id": secretaria_id,
            "unidades": unidades,
        },
    )


@login_required
def relatorio_estoque(request):
    categoria_id = request.GET.get("categoria_id", "")
    apenas_criticos = request.GET.get("apenas_criticos", "")

    produtos = Produto.objects.select_related("categoria").order_by("categoria__nome", "nome")
    if categoria_id:
        produtos = produtos.filter(categoria_id=categoria_id)
    if apenas_criticos:
        produtos = produtos.filter(estoque_atual__lt=F("estoque_minimo"))

    totais = produtos.aggregate(
        total_estoque=Sum("estoque_atual"),
        total_reservado=Sum("estoque_reservado"),
        total_minimo=Sum("estoque_minimo"),
    )
    total_estoque = _to_decimal(totais.get("total_estoque"))
    total_reservado = _to_decimal(totais.get("total_reservado"))

    export = request.GET.get("export", "")
    if export in ("csv", "xlsx"):
        headers = ["Produto", "Categoria", "Unid.", "Físico", "Reservado", "Disponível", "Mínimo"]
        rows = [
            [
                p.nome, p.categoria.nome, p.unidade_medida,
                p.estoque_atual, p.estoque_reservado, p.estoque_disponivel, p.estoque_minimo,
            ]
            for p in produtos
        ]
        if export == "csv":
            return _export_csv("relatorio_estoque.csv", headers, rows)
        return _export_xlsx("relatorio_estoque.xlsx", [("Estoque", headers, rows)])

    from .models import Categoria
    categorias = Categoria.objects.order_by("nome")
    return render(
        request,
        "estoque/relatorio_estoque.html",
        {
            "produtos": produtos,
            "total_produtos": produtos.count(),
            "produtos_criticos": Produto.objects.filter(estoque_atual__lt=F("estoque_minimo")).count(),
            "total_estoque": total_estoque,
            "total_reservado": total_reservado,
            "total_disponivel": total_estoque - total_reservado,
            "total_minimo": _to_decimal(totais.get("total_minimo")),
            "categorias": categorias,
            "categoria_id": categoria_id,
            "apenas_criticos": apenas_criticos,
        },
    )


@login_required
def relatorio_pedidos(request):
    data_de = request.GET.get("data_de", "")
    data_ate = request.GET.get("data_ate", "")
    status_filtro = request.GET.get("status", "")
    secretaria_id = request.GET.get("secretaria_id", "")

    pedidos = Pedido.objects.select_related("secretaria").order_by("-data_pedido")
    if data_de:
        pedidos = pedidos.filter(data_pedido__date__gte=data_de)
    if data_ate:
        pedidos = pedidos.filter(data_pedido__date__lte=data_ate)
    if status_filtro:
        pedidos = pedidos.filter(status=status_filtro)
    if secretaria_id:
        pedidos = pedidos.filter(secretaria_id=secretaria_id)

    resumo_status = (
        pedidos.values("status").annotate(total=Count("id")).order_by("status")
    )
    status_labels = dict(Pedido.STATUS_CHOICES)
    resumo_status_display = [
        {
            "status": item["status"],
            "status_label": status_labels.get(item["status"], item["status"]),
            "total": item["total"],
        }
        for item in resumo_status
    ]
    totais_itens = ItemPedido.objects.filter(pedido__in=pedidos).aggregate(
        total_solicitado=Sum("quantidade"),
        total_atendido=Sum("quantidade_atendida"),
    )

    export = request.GET.get("export", "")
    if export in ("csv", "xlsx"):
        headers = ["ID", "Data", "Secretaria", "Status", "Endereço", "Observação"]
        rows = [
            [
                p.pk,
                p.data_pedido.strftime("%d/%m/%Y"),
                p.secretaria.nome if p.secretaria else "",
                p.get_status_display(),
                p.endereco_entrega,
                p.observacoes,
            ]
            for p in pedidos
        ]
        if export == "csv":
            return _export_csv("relatorio_pedidos.csv", headers, rows)
        return _export_xlsx("relatorio_pedidos.xlsx", [("Pedidos", headers, rows)])

    unidades = Unidade.objects.order_by("nome")
    return render(
        request,
        "estoque/relatorio_pedidos.html",
        {
            "pedidos": pedidos[:100],
            "resumo_status": resumo_status_display,
            "total_pedidos": pedidos.count(),
            "total_solicitado": _to_decimal(totais_itens.get("total_solicitado")),
            "total_atendido": _to_decimal(totais_itens.get("total_atendido")),
            "data_de": data_de,
            "data_ate": data_ate,
            "status_filtro": status_filtro,
            "secretaria_id": secretaria_id,
            "unidades": unidades,
            "status_choices": Pedido.STATUS_CHOICES,
        },
    )


@_tem_papel("Almoxarife", "Comprador", "Administrador")
def fornecedor_update(request, pk):
    fornecedor = get_object_or_404(Fornecedor, pk=pk)
    form = FornecedorForm(request.POST or None, instance=fornecedor)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Fornecedor atualizado com sucesso!")
        return redirect("fornecedor_list")
    return render(
        request,
        "estoque/fornecedor_form.html",
        {"form": form, "title": "Editar Fornecedor"},
    )


@_tem_papel("Almoxarife", "Administrador")
def fornecedor_delete(request, pk):
    fornecedor = get_object_or_404(Fornecedor, pk=pk)
    if request.method == "POST":
        try:
            fornecedor.delete()
            messages.success(request, "Fornecedor excluído com sucesso!")
        except ProtectedError:
            messages.error(
                request,
                "Não foi possível excluir o fornecedor porque ele está vinculado a entradas ou pedidos.",
            )
    return redirect("fornecedor_list")


@_tem_papel("Almoxarife", "Administrador")
def produto_delete(request, pk):
    produto = get_object_or_404(Produto, pk=pk)
    if request.method == "POST":
        try:
            produto.delete()
            messages.success(request, "Produto excluído com sucesso!")
        except ProtectedError:
            messages.error(
                request,
                "Não foi possível excluir o produto porque ele está vinculado a entradas ou pedidos.",
            )
    return redirect("produto_list")


@_tem_papel("Almoxarife", "Comprador", "Administrador")
def importar_licitacao(request):
    preview = None
    erros = []
    licitacao_nome = ""

    if request.method == "POST":
        arquivo = request.FILES.get("arquivo")
        licitacao_nome = request.POST.get("licitacao_nome", "").strip()
        confirmar = request.POST.get("confirmar", "")

        if not arquivo:
            messages.error(request, "Selecione um arquivo XLSX para importar.")
            return render(request, "estoque/importar_licitacao.html", {"licitacao_nome": licitacao_nome})

        if not arquivo.name.endswith((".xlsx", ".xls")):
            messages.error(request, "Apenas arquivos Excel (.xlsx ou .xls) são aceitos.")
            return render(request, "estoque/importar_licitacao.html", {"licitacao_nome": licitacao_nome})

        try:
            from openpyxl import load_workbook
            wb = load_workbook(arquivo, read_only=True, data_only=True)
            ws = wb.active

            linhas = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # pula cabeçalho
                nome_produto = str(row[0]).strip() if row[0] else ""
                quantidade_raw = row[1] if len(row) > 1 else None
                preco_raw = row[2] if len(row) > 2 else None

                if not nome_produto:
                    continue

                try:
                    quantidade = Decimal(str(quantidade_raw)) if quantidade_raw else Decimal("0")
                    preco = Decimal(str(preco_raw)) if preco_raw else Decimal("0")
                except Exception:
                    erros.append(f"Linha {i + 1}: quantidade ou preço inválidos.")
                    continue

                produto = Produto.objects.filter(nome__iexact=nome_produto).first()
                linhas.append({
                    "nome": nome_produto,
                    "produto": produto,
                    "quantidade": quantidade,
                    "preco": preco,
                    "encontrado": produto is not None,
                })

            wb.close()

            if confirmar and not erros:
                importados = 0
                for linha in linhas:
                    if not linha["encontrado"]:
                        continue
                    produto = linha["produto"]
                    produto.estoque_atual += linha["quantidade"]
                    produto.save()
                    importados += 1
                messages.success(request, f"Importação concluída: {importados} produto(s) com estoque atualizado.")
                return redirect("importar_licitacao")

            preview = linhas

        except Exception as e:
            messages.error(request, f"Erro ao processar o arquivo: {e}")

    return render(
        request,
        "estoque/importar_licitacao.html",
        {
            "preview": preview,
            "erros": erros,
            "licitacao_nome": licitacao_nome,
        },
    )


@_tem_papel("Almoxarife", "Administrador")
def entrada_update(request, pk):
    entrada = get_object_or_404(Entrada, pk=pk)
    form = EntradaForm(request.POST or None, instance=entrada)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Entrada atualizada com sucesso!")
        return redirect("entrada_list")
    return render(
        request,
        "estoque/entrada_form.html",
        {
            "entrada_form": form,
            "formset": None,
            "is_edit": True,
            "entrada": entrada,
        },
    )


@_tem_papel("Almoxarife", "Administrador")
def entrada_delete(request, pk):
    entrada = get_object_or_404(Entrada, pk=pk)
    if request.method == "POST":
        for item in entrada.itens.select_related("produto"):
            produto = item.produto
            produto.estoque_atual -= item.quantidade
            produto.save()
        entrada.delete()
        messages.success(request, "Entrada excluída e estoque estornado com sucesso!")
    return redirect("entrada_list")


@login_required
def profile(request):
    perfil_usuario, _ = PerfilUsuario.objects.get_or_create(user=request.user)
    perfil_form = PerfilUsuarioForm(request.POST or None, instance=request.user)
    tema_form = PerfilTemaForm(request.POST or None, instance=perfil_usuario)
    senha_form = PasswordChangeForm(user=request.user, data=request.POST or None)

    if request.method == "POST":
        acao = request.POST.get("acao")
        if acao == "dados":
            if perfil_form.is_valid() and tema_form.is_valid():
                perfil_form.save()
                tema_form.save()
                messages.success(request, "Perfil atualizado com sucesso.")
                return redirect("profile")
            messages.error(request, "Revise os dados do perfil e tente novamente.")

        elif acao == "senha":
            if senha_form.is_valid():
                user = senha_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Senha alterada com sucesso.")
                return redirect("profile")
            messages.error(request, "Não foi possível alterar a senha.")

    return render(
        request,
        "registration/profile.html",
        {
            "perfil_form": perfil_form,
            "tema_form": tema_form,
            "senha_form": senha_form,
        },
    )


# --- PDF ---
def _render_pdf(template_name, context, filename):
    """Renderiza um template HTML como PDF usando WeasyPrint."""
    from weasyprint import HTML
    from django.template.loader import render_to_string
    html_string = render_to_string(template_name, context)
    pdf_file = HTML(string=html_string, base_url=None).write_pdf()
    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


@login_required
def pedido_pdf(request, pk):
    pedido = get_object_or_404(
        Pedido.objects.select_related("secretaria", "solicitante").prefetch_related(
            "itens__produto"
        ),
        pk=pk,
    )
    return _render_pdf(
        "estoque/pedido_pdf.html",
        {"pedido": pedido},
        f"pedido_{pedido.pk}.pdf",
    )


@login_required
def produto_ficha_pdf(request, pk):
    produto = get_object_or_404(Produto.objects.select_related("categoria"), pk=pk)
    entradas = (
        ItemEntrada.objects.filter(produto=produto)
        .select_related("entrada")
        .order_by("-entrada__data_entrada")[:50]
    )
    saidas = (
        ItemPedido.objects.filter(produto=produto)
        .select_related("pedido__secretaria")
        .order_by("-pedido__data_pedido")[:50]
    )
    return _render_pdf(
        "estoque/produto_ficha_pdf.html",
        {"produto": produto, "entradas": entradas, "saidas": saidas},
        f"ficha_{produto.pk}.pdf",
    )
