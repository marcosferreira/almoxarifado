from decimal import Decimal
from typing import Any

from django.http import HttpRequest, HttpResponse
from django.db import transaction
from django.contrib import messages
from ._base import (
    render, redirect,
    login_required, _tem_papel,
    Produto,
)


UNIDADES_VALIDAS = {"UN", "CX", "KG", "LT", "PCT", "RM"}


def _normalizar_unidade(valor: str) -> str:
    valor = str(valor or "").strip().upper()
    if valor in UNIDADES_VALIDAS:
        return valor
    return "UN"


def _parse_decimal(valor: Any) -> Decimal:
    try:
        return Decimal(str(valor or "0"))
    except Exception:
        return Decimal("0")


@_tem_papel("Almoxarife", "Comprador", "Administrador")
def importar_produtos(request: HttpRequest) -> HttpResponse:
    preview: list[dict[str, Any]] = []
    erros: list[str] = []
    nome_importacao = ""

    if request.method == "POST":
        confirmar = request.POST.get("confirmar", "")

        if confirmar:
            session_data = request.session.pop("importar_produtos", None)
            if not session_data:
                messages.error(request, "Sessão expirada. Faça o upload novamente.")
                return render(request, "estoque/importar_produtos.html", {"nome_importacao": nome_importacao})

            importados = 0
            pulados = 0
            with transaction.atomic():
                for item in session_data:
                    nome = item["nome"]
                    categoria_nome = item["categoria"]
                    unidade = item["unidade"]
                    estoque_min = Decimal(item["estoque_minimo"])

                    if not nome:
                        continue

                    from ..models import Categoria
                    categoria, _ = Categoria.objects.get_or_create(nome=categoria_nome)

                    produto, created = Produto.objects.get_or_create(
                        nome__iexact=nome,
                        defaults={
                            "nome": nome,
                            "categoria": categoria,
                            "unidade_medida": unidade,
                            "estoque_minimo": estoque_min,
                        },
                    )
                    if created:
                        importados += 1
                    else:
                        pulados += 1

            messages.success(
                request,
                f"Importação concluída: {importados} produto(s) criado(s), {pulados} já existente(s).",
            )
            return redirect("importar_produtos")

        arquivo = request.FILES.get("arquivo")
        if not arquivo:
            messages.error(request, "Selecione um arquivo XLSX para importar.")
            return render(request, "estoque/importar_produtos.html", {"nome_importacao": nome_importacao})

        if not arquivo.name.lower().endswith(".xlsx"):
            messages.error(request, "Apenas arquivos Excel (.xlsx) são aceitos.")
            return render(request, "estoque/importar_produtos.html", {"nome_importacao": nome_importacao})

        try:
            from openpyxl import load_workbook
            wb = load_workbook(arquivo, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                messages.error(request, "Planilha sem abas.")
                wb.close()
                return render(request, "estoque/importar_produtos.html", {"nome_importacao": nome_importacao})

            session_data: list[dict[str, Any]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # header row

                nome = str(row[0] or "").strip() if len(row) > 0 else ""
                if not nome:
                    continue

                categoria_nome = str(row[1] or "Importação").strip() if len(row) > 1 else "Importação"
                if not categoria_nome:
                    categoria_nome = "Importação"

                unidade = _normalizar_unidade(str(row[2] or "")) if len(row) > 2 else "UN"
                estoque_min = _parse_decimal(row[3] if len(row) > 3 else None)

                ja_existe = Produto.objects.filter(nome__iexact=nome).exists()

                preview.append({
                    "nome": nome,
                    "categoria": categoria_nome,
                    "unidade": unidade,
                    "estoque_minimo": float(estoque_min),
                    "encontrado": ja_existe,
                })

                if not ja_existe:
                    session_data.append({
                        "nome": nome,
                        "categoria": categoria_nome,
                        "unidade": unidade,
                        "estoque_minimo": str(estoque_min),
                    })

            wb.close()

            if not preview:
                messages.warning(request, "Nenhum produto encontrado na planilha (apenas cabeçalhos?).")
            else:
                request.session["importar_produtos"] = session_data

        except Exception as e:
            messages.error(request, f"Erro ao processar o arquivo: {e}")

    return render(
        request,
        "estoque/importar_produtos.html",
        {"preview": preview, "erros": erros, "nome_importacao": nome_importacao},
    )
