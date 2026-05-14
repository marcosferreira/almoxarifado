import re
from typing import Any

from django.http import HttpRequest, HttpResponse
from django.utils import timezone
from ._base import (
    render, redirect,
    messages, Decimal, F, Sum, Count,
    login_required, _tem_papel, _to_decimal, _export_csv, _export_xlsx,
    Produto, Pedido, ItemPedido, Entrada, ItemEntrada, Fornecedor, Unidade,
)


@login_required
def relatorios(request: HttpRequest) -> HttpResponse:
    return redirect("relatorio_movimento")


@login_required
def relatorio_movimento(request: HttpRequest) -> HttpResponse:
    data_de = request.GET.get("data_de", "")
    data_ate = request.GET.get("data_ate", "")
    secretaria_id = request.GET.get("secretaria_id", "")

    consumo_qs = ItemPedido.objects.filter(pedido__status="ENTREGUE")
    if data_de:
        consumo_qs = consumo_qs.filter(pedido__data_pedido__date__gte=data_de)
    if data_ate:
        consumo_qs = consumo_qs.filter(pedido__data_pedido__date__lte=data_ate)
    if secretaria_id:
        consumo_qs = consumo_qs.filter(pedido__secretaria_id=secretaria_id)

    consumo_secretaria = (
        consumo_qs
        .values("pedido__secretaria__nome")
        .annotate(total_itens=Sum("quantidade"))
        .order_by("-total_itens")
    )

    entradas_qs = ItemEntrada.objects.all()
    if data_de:
        entradas_qs = entradas_qs.filter(entrada__data_entrada__gte=data_de)
    if data_ate:
        entradas_qs = entradas_qs.filter(entrada__data_entrada__lte=data_ate)

    entradas_por_grupo = (
        entradas_qs
        .values("produto__categoria__nome")
        .annotate(total_quantidade=Sum("quantidade"))
        .order_by("-total_quantidade")
    )

    export = request.GET.get("export", "")
    if export == "csv":
        rows = [
            [r["pedido__secretaria__nome"] or "Sem secretaria", r["total_itens"]]
            for r in consumo_secretaria
        ]
        return _export_csv("consumo_secretaria.csv", ["Secretaria", "Total Itens"], rows)
    if export == "xlsx":
        rows_consumo = [
            [r["pedido__secretaria__nome"] or "Sem secretaria", r["total_itens"]]
            for r in consumo_secretaria
        ]
        rows_entradas = [
            [r["produto__categoria__nome"] or "Sem categoria", r["total_quantidade"]]
            for r in entradas_por_grupo
        ]
        return _export_xlsx("relatorio_movimento.xlsx", [
            ("Consumo por Secretaria", ["Secretaria", "Total Itens"], rows_consumo),
            ("Entradas por Categoria", ["Categoria", "Total Quantidade"], rows_entradas),
        ])

    unidades = Unidade.objects.order_by("nome")
    return render(
        request,
        "estoque/relatorios.html",
        {
            "consumo_secretaria": consumo_secretaria,
            "entradas_por_grupo": entradas_por_grupo,
            "data_de": data_de,
            "data_ate": data_ate,
            "secretaria_id": secretaria_id,
            "unidades": unidades,
        },
    )


@login_required
def relatorio_estoque(request: HttpRequest) -> HttpResponse:
    categoria_id = request.GET.get("categoria_id", "")
    apenas_criticos = request.GET.get("apenas_criticos", "")

    produtos = Produto.objects.select_related("categoria").order_by("categoria__nome", "nome")
    if categoria_id:
        produtos = produtos.filter(categoria_id=categoria_id)
    if apenas_criticos:
        produtos = produtos.filter(estoque_atual__lt=F("estoque_minimo"))

    totais = produtos.aggregate(
        total_estoque=Sum("estoque_atual"),
        total_reservado=Sum("estoque_reservado"),
        total_minimo=Sum("estoque_minimo"),
    )
    total_estoque = _to_decimal(totais.get("total_estoque"))
    total_reservado = _to_decimal(totais.get("total_reservado"))

    export = request.GET.get("export", "")
    if export in ("csv", "xlsx"):
        headers = ["Produto", "Categoria", "Unid.", "Físico", "Reservado", "Disponível", "Mínimo"]
        rows = [
            [
                p.nome, p.categoria.nome, p.unidade_medida,
                p.estoque_atual, p.estoque_reservado, p.estoque_disponivel, p.estoque_minimo,
            ]
            for p in produtos
        ]
        if export == "csv":
            return _export_csv("relatorio_estoque.csv", headers, rows)
        return _export_xlsx("relatorio_estoque.xlsx", [("Estoque", headers, rows)])

    from ..models import Categoria
    categorias = Categoria.objects.order_by("nome")
    return render(
        request,
        "estoque/relatorio_estoque.html",
        {
            "produtos": produtos,
            "total_produtos": produtos.count(),
            "produtos_criticos": Produto.objects.filter(estoque_atual__lt=F("estoque_minimo")).count(),
            "total_estoque": total_estoque,
            "total_reservado": total_reservado,
            "total_disponivel": total_estoque - total_reservado,
            "total_minimo": _to_decimal(totais.get("total_minimo")),
            "categorias": categorias,
            "categoria_id": categoria_id,
            "apenas_criticos": apenas_criticos,
        },
    )


@login_required
def relatorio_pedidos(request: HttpRequest) -> HttpResponse:
    data_de = request.GET.get("data_de", "")
    data_ate = request.GET.get("data_ate", "")
    status_filtro = request.GET.get("status", "")
    secretaria_id = request.GET.get("secretaria_id", "")

    pedidos = Pedido.objects.select_related("secretaria").order_by("-data_pedido")
    if data_de:
        pedidos = pedidos.filter(data_pedido__date__gte=data_de)
    if data_ate:
        pedidos = pedidos.filter(data_pedido__date__lte=data_ate)
    if status_filtro:
        pedidos = pedidos.filter(status=status_filtro)
    if secretaria_id:
        pedidos = pedidos.filter(secretaria_id=secretaria_id)

    resumo_status = (
        pedidos.values("status").annotate(total=Count("id")).order_by("status")
    )
    status_labels = dict(Pedido.STATUS_CHOICES)
    resumo_status_display = [
        {
            "status": item["status"],
            "status_label": status_labels.get(item["status"], item["status"]),
            "total": item["total"],
        }
        for item in resumo_status
    ]
    totais_itens = ItemPedido.objects.filter(pedido__in=pedidos).aggregate(
        total_solicitado=Sum("quantidade"),
        total_atendido=Sum("quantidade_atendida"),
    )

    export = request.GET.get("export", "")
    if export in ("csv", "xlsx"):
        headers = ["ID", "Data", "Secretaria", "Status", "Endereço", "Observação"]
        rows = [
            [
                p.pk,
                p.data_pedido.strftime("%d/%m/%Y"),
                p.secretaria.nome if p.secretaria else "",
                p.get_status_display(),
                p.endereco_entrega,
                p.observacoes,
            ]
            for p in pedidos
        ]
        if export == "csv":
            return _export_csv("relatorio_pedidos.csv", headers, rows)
        return _export_xlsx("relatorio_pedidos.xlsx", [("Pedidos", headers, rows)])

    unidades = Unidade.objects.order_by("nome")
    return render(
        request,
        "estoque/relatorio_pedidos.html",
        {
            "pedidos": pedidos[:100],
            "resumo_status": resumo_status_display,
            "total_pedidos": pedidos.count(),
            "total_solicitado": _to_decimal(totais_itens.get("total_solicitado")),
            "total_atendido": _to_decimal(totais_itens.get("total_atendido")),
            "data_de": data_de,
            "data_ate": data_ate,
            "status_filtro": status_filtro,
            "secretaria_id": secretaria_id,
            "unidades": unidades,
            "status_choices": Pedido.STATUS_CHOICES,
        },
    )


def _parse_proponente(cell_text: str) -> tuple[str, str]:
    """Extrai nome e CNPJ do texto da célula A1 de cada aba."""
    lines = [l.strip() for l in str(cell_text).split("\n") if l.strip()]
    nome = lines[0] if lines else ""
    # Remove prefixo "PROPONENTE:" se presente
    nome = re.sub(r'^PROPONENTE\s*:\s*', '', nome, flags=re.IGNORECASE).strip()
    cnpj = ""
    cnpj_match = re.search(r'(\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})', str(cell_text))
    if cnpj_match:
        cnpj = cnpj_match.group(1)
    return nome, cnpj


def _normalizar_cnpj(cnpj_raw: str) -> str:
    digits = re.sub(r"\D", "", str(cnpj_raw or ""))
    if len(digits) != 14:
        return ""
    return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:14]}"


def _normalizar_texto_limite(valor: str, limite: int = 200) -> str:
    texto = re.sub(r"\s+", " ", str(valor or "")).strip()
    if len(texto) <= limite:
        return texto
    return texto[:limite].rstrip()


def _obter_ou_criar_fornecedor(nome_forn, cnpj_forn):
    nome_forn = _normalizar_texto_limite(nome_forn, limite=200) or "Fornecedor sem nome"
    cnpj_normalizado = _normalizar_cnpj(cnpj_forn)
    if cnpj_normalizado:
        fornecedor, _ = Fornecedor.objects.get_or_create(
            cnpj=cnpj_normalizado,
            defaults={"nome_fantasia": nome_forn, "razao_social": nome_forn},
        )
        return fornecedor

    fornecedor = Fornecedor.objects.filter(nome_fantasia__iexact=nome_forn).first()
    if fornecedor:
        return fornecedor

    sequencia = Fornecedor.objects.count() + 1
    while True:
        bloco = sequencia // 100
        sufixo = sequencia % 100
        cnpj_placeholder = f"99.999.999/{bloco:04d}-{sufixo:02d}"
        if not Fornecedor.objects.filter(cnpj=cnpj_placeholder).exists():
            break
        sequencia += 1

    return Fornecedor.objects.create(
        nome_fantasia=nome_forn,
        razao_social=nome_forn,
        cnpj=cnpj_placeholder,
    )


def _obter_ou_criar_produto(nome_produto):
    nome_produto = _normalizar_texto_limite(nome_produto, limite=200)
    if not nome_produto:
        nome_produto = "Produto sem descrição"

    produto = Produto.objects.filter(nome__iexact=nome_produto).first()
    if produto:
        return produto, False

    from ..models import Categoria

    categoria_importacao, _ = Categoria.objects.get_or_create(nome="Importação Licitação")
    produto = Produto.objects.create(
        nome=nome_produto,
        categoria=categoria_importacao,
        unidade_medida="UN",
    )
    return produto, True


def _executar_importacao(proponentes_sessao: list[dict[str, Any]], licitacao_nome: str, criado_por: Any = None) -> tuple[int, int]:
    """Cria Fornecedor, Entrada e ItemEntrada para cada proponente. Retorna o total importado."""
    importados = 0
    produtos_criados = 0
    hoje = timezone.now().date()
    licitacao_nome = _normalizar_texto_limite(licitacao_nome, limite=200)
    for prop in proponentes_sessao:
        nome_forn = prop["nome"]
        cnpj_forn = prop["cnpj"]
        itens = prop["itens"]
        if not itens:
            continue

        fornecedor = _obter_ou_criar_fornecedor(nome_forn, cnpj_forn)

        entrada = Entrada.objects.create(
            fornecedor=fornecedor,
            data_entrada=hoje,
            licitacao=licitacao_nome,
            criado_por=criado_por,
        )

        for item in itens:
            try:
                produto_id = item.get("produto_id")
                if produto_id:
                    produto = Produto.objects.get(pk=produto_id)
                else:
                    produto, criado = _obter_ou_criar_produto(item["produto_nome"])
                    if criado:
                        produtos_criados += 1
                ItemEntrada.objects.create(
                    entrada=entrada,
                    produto=produto,
                    quantidade=Decimal(item["quantidade"]),
                    preco_unitario=Decimal(item["preco"]),
                )
                importados += 1
            except Produto.DoesNotExist:
                pass

    return importados, produtos_criados


@_tem_papel("Almoxarife", "Comprador", "Administrador")
def importar_licitacao(request: HttpRequest) -> HttpResponse:
    preview = None
    erros = []
    licitacao_nome = ""

    if request.method == "POST":
        licitacao_nome = request.POST.get("licitacao_nome", "").strip()
        confirmar = request.POST.get("confirmar", "")

        # Confirmação usa os dados já gravados na sessão (não precisa do arquivo novamente)
        if confirmar:
            proponentes_sessao = request.session.pop("licitacao_importar", None)
            if not proponentes_sessao:
                messages.error(request, "Sessão expirada ou sem dados para importar. Faça o upload novamente.")
                return render(request, "estoque/importar_licitacao.html", {"licitacao_nome": licitacao_nome})
            importados, produtos_criados = _executar_importacao(proponentes_sessao, licitacao_nome, criado_por=request.user)
            messages.success(
                request,
                f"Importação concluída: {importados} item(ns) registrado(s) em entrada de estoque. Produtos novos criados: {produtos_criados}.",
            )
            return redirect("importar_licitacao")

        arquivo = request.FILES.get("arquivo")
        importar_direto = request.POST.get("importar_direto", "")

        if not arquivo:
            messages.error(request, "Selecione um arquivo XLSX para importar.")
            return render(request, "estoque/importar_licitacao.html", {"licitacao_nome": licitacao_nome})

        if not arquivo.name.lower().endswith(".xlsx"):
            messages.error(request, "Apenas arquivos Excel (.xlsx) são aceitos.")
            return render(request, "estoque/importar_licitacao.html", {"licitacao_nome": licitacao_nome})

        try:
            from openpyxl import load_workbook
            wb = load_workbook(arquivo, read_only=True, data_only=True)

            # Estrutura: { aba_title: {"nome": str, "cnpj": str, "itens": [...]} }
            proponentes = {}
            linhas = []

            for ws in wb.worksheets:
                proponente_nome = ws.title
                proponente_cnpj = ""

                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0 and row and row[0]:
                        proponente_nome, proponente_cnpj = _parse_proponente(str(row[0]))
                    # Pula as 3 primeiras linhas (proponente, mensagem, cabeçalhos)
                    if i < 3:
                        continue

                    col_a = row[0] if row else None
                    col_f = row[5] if len(row) > 5 else None

                    # Linha de totais: col A vazia e col F com valor → fim da lista de produtos
                    if not col_a and col_f is not None:
                        break

                    nome_produto = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    nome_produto_normalizado = _normalizar_texto_limite(nome_produto, limite=200)
                    quantidade_raw = row[3] if len(row) > 3 else None
                    preco_raw = row[4] if len(row) > 4 else None

                    if not nome_produto_normalizado:
                        continue

                    try:
                        quantidade = Decimal(str(quantidade_raw)) if quantidade_raw else Decimal("0")
                        preco = Decimal(str(preco_raw)) if preco_raw else Decimal("0")
                    except Exception:
                        erros.append(f"Aba '{ws.title}', linha {i + 1}: quantidade ou preço inválidos.")
                        continue

                    produto = Produto.objects.filter(nome__iexact=nome_produto_normalizado).first()
                    linhas.append({
                        "proponente": proponente_nome,
                        "nome": nome_produto,
                        "produto": produto,
                        "quantidade": quantidade,
                        "preco": preco,
                        "encontrado": produto is not None,
                        "criara_novo": produto is None,
                    })

                    prop = proponentes.setdefault(ws.title, {
                        "nome": proponente_nome,
                        "cnpj": proponente_cnpj,
                        "itens": [],
                    })
                    prop["itens"].append({
                        "produto_id": produto.pk if produto else None,
                        "produto_nome": nome_produto_normalizado,
                        "quantidade": str(quantidade),
                        "preco": str(preco),
                    })

            wb.close()

            # Salva estrutura por proponente na sessão
            request.session["licitacao_importar"] = list(proponentes.values())

            if importar_direto and not erros:
                proponentes_sessao = request.session.pop("licitacao_importar", [])
                importados, produtos_criados = _executar_importacao(proponentes_sessao, licitacao_nome, criado_por=request.user)
                messages.success(
                    request,
                    f"Importação concluída: {importados} item(ns) registrado(s) em entrada de estoque. Produtos novos criados: {produtos_criados}.",
                )
                return redirect("importar_licitacao")

            preview = linhas

        except Exception as e:
            messages.error(request, f"Erro ao processar o arquivo: {e}")

    return render(
        request,
        "estoque/importar_licitacao.html",
        {"preview": preview, "erros": erros, "licitacao_nome": licitacao_nome},
    )
